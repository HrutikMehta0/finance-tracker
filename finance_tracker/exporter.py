"""
Transaction export utilities.

Exports transactions to various formats (Excel, future: Google Sheets, CSV).
"""
import logging
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .models import Transaction
from .constants import TRANSACTION_SHEET_NAME, TRANSACTION_SHEET_COLUMNS
from .logging_config import get_logger
from .deduplicator import generate_transaction_hash
from .normalizer import normalize_description

logger = get_logger(__name__)


class Exporter:
    """Handles transaction export to various formats."""

    @staticmethod
    def get_existing_transaction_ids(workbook_path: str) -> set:
        """Return a set of existing transaction IDs from the workbook transaction sheet."""
        wb_path = Path(workbook_path)
        if not wb_path.exists():
            return set()

        wb = load_workbook(wb_path)
        if TRANSACTION_SHEET_NAME not in wb.sheetnames:
            return set()

        ws = wb[TRANSACTION_SHEET_NAME]
        # Find Transaction ID column index
        header = [cell.value if cell.value is not None else "" for cell in ws[1]]
        txid_idx = None
        for idx, h in enumerate(header, start=1):
            if isinstance(h, str) and "transaction" in h.lower() and "id" in h.lower():
                txid_idx = idx
                break

        ids = set()
        if txid_idx is not None:
            for row in ws.iter_rows(min_row=2, min_col=txid_idx, max_col=txid_idx, values_only=True):
                val = row[0]
                if val:
                    ids.add(str(val))
            return ids

        # If Transaction ID column missing, compute stable hashes from existing rows
        # Attempt to find date, description, amount, account columns
        col_map = {str(h).lower(): idx + 1 for idx, h in enumerate(header)}
        date_idx = None
        desc_idx = None
        amount_idx = None
        account_idx = None

        for name, idx in col_map.items():
            if "date" in name and date_idx is None:
                date_idx = idx
            if "transaction name" in name or "description" in name or "merchant" in name:
                desc_idx = idx
            if "amount" in name:
                amount_idx = idx
            if "spent" in name or "account" in name or "spent from" in name:
                account_idx = idx

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                date_val = row[date_idx - 1] if date_idx and len(row) >= date_idx else None
                desc_val = row[desc_idx - 1] if desc_idx and len(row) >= desc_idx else ""
                amount_val = row[amount_idx - 1] if amount_idx and len(row) >= amount_idx else 0.0
                account_val = row[account_idx - 1] if account_idx and len(row) >= account_idx else ""

                if date_val is None:
                    continue

                # Normalize values and compute hash
                txn_hash = generate_transaction_hash(date_val, str(desc_val), float(amount_val or 0.0), str(account_val or ""))
                ids.add(str(txn_hash))
            except Exception:
                continue

        return ids

    @staticmethod
    def filter_new_transactions(transactions: List[Transaction], existing_ids: set) -> List[Transaction]:
        """Return subset of transactions whose transaction_id is not in existing_ids."""
        new = [t for t in transactions if (t.transaction_id is None) or (str(t.transaction_id) not in existing_ids)]
        return new

    @staticmethod
    def to_excel(
        transactions: List[Transaction],
        output_path: str,
        include_id: bool = False,
    ) -> None:
        """
        Export transactions to an Excel file.

        Args:
            transactions: List of Transaction objects to export.
            output_path: Path for output Excel file.
            include_id: If True, includes transaction_id column.

        Raises:
            ValueError: If transactions list is empty.
            IOError: If file cannot be written.
        """
        if not transactions:
            raise ValueError("No transactions to export")

        try:
            # Convert transactions to DataFrame
            data = []
            for txn in transactions:
                row = {
                    "Date": txn.date,
                    "Description": txn.name,
                    "Amount": txn.amount,
                    "Category": txn.category,
                    "Account": txn.spent_from,
                    "Notes": txn.notes,
                }
                if include_id:
                    row["Transaction ID"] = txn.transaction_id
                data.append(row)

            df = pd.DataFrame(data)

            # Ensure date column is properly formatted
            if "Date" in df.columns:
                df["Date"] = pd.to_datetime(df["Date"])

            # Write to Excel by appending rows to existing workbook to preserve
            # formulas, charts, and other sheets. Use openpyxl to avoid full
            # workbook overwrite.
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)

            if not output_file.exists():
                # Create a new workbook with header row
                wb = load_workbook(filename=None) if False else None
                # Fallback: use pandas to create a minimal workbook then reopen
                df_head = pd.DataFrame(columns=TRANSACTION_SHEET_COLUMNS)
                with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                    df_head.to_excel(writer, sheet_name=TRANSACTION_SHEET_NAME, index=False)

            # Load workbook and append rows
            wb = load_workbook(output_path)
            if TRANSACTION_SHEET_NAME not in wb.sheetnames:
                # Create the transactions sheet with header
                ws = wb.create_sheet(TRANSACTION_SHEET_NAME)
                for idx, col in enumerate(TRANSACTION_SHEET_COLUMNS, start=1):
                    ws.cell(row=1, column=idx, value=col)
            else:
                ws = wb[TRANSACTION_SHEET_NAME]

            # Read existing transaction ids and filter incoming transactions
            existing_ids = Exporter.get_existing_transaction_ids(str(output_file))
            new_txns = Exporter.filter_new_transactions(transactions, existing_ids)

            duplicates_detected = len(transactions) - len(new_txns)
            if duplicates_detected > 0:
                logger.info(f"Detected {duplicates_detected} duplicate transactions; they will be skipped")

            if not new_txns:
                logger.info("No new transactions to append")
                return

            # Determine header and transaction id column
            header = [cell.value if cell.value is not None else "" for cell in ws[1]]
            try:
                txid_col_idx = header.index("Transaction ID") + 1
            except ValueError:
                # Add Transaction ID column at the end
                txid_col_idx = len(header) + 1
                ws.cell(row=1, column=txid_col_idx, value="Transaction ID")

            # Determine next empty row
            next_row = ws.max_row + 1

            for txn in new_txns:
                values = [
                    txn.date,
                    txn.name,
                    txn.category,
                    txn.spent_from,
                    txn.amount,
                    txn.notes,
                ]
                for col_idx, value in enumerate(values, start=1):
                    ws.cell(row=next_row, column=col_idx, value=value)
                # Write transaction id
                ws.cell(row=next_row, column=txid_col_idx, value=txn.transaction_id)
                next_row += 1

            wb.save(output_path)
            logger.info(f"Appended {len(new_txns)} new transactions to {output_path}")
            logger.info(f"Workbook now has {ws.max_row} rows in sheet '{TRANSACTION_SHEET_NAME}'")

        except Exception as e:
            logger.error(f"Failed to export to Excel: {e}")
            raise

    @staticmethod
    def to_csv(
        transactions: List[Transaction],
        output_path: str,
    ) -> None:
        """
        Export transactions to a CSV file.

        Args:
            transactions: List of Transaction objects to export.
            output_path: Path for output CSV file.

        Raises:
            ValueError: If transactions list is empty.
            IOError: If file cannot be written.
        """
        if not transactions:
            raise ValueError("No transactions to export")

        try:
            data = [
                {
                    "Date": txn.date,
                    "Description": txn.name,
                    "Amount": txn.amount,
                    "Category": txn.category,
                    "Account": txn.spent_from,
                    "Notes": txn.notes,
                    "Transaction ID": txn.transaction_id,
                }
                for txn in transactions
            ]

            df = pd.DataFrame(data)
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)

            df.to_csv(output_path, index=False)
            logger.info(f"Exported {len(transactions)} transactions to {output_path}")

        except Exception as e:
            logger.error(f"Failed to export to CSV: {e}")
            raise

    @staticmethod
    def to_dataframe(transactions: List[Transaction]) -> pd.DataFrame:
        """
        Convert transactions to a pandas DataFrame.

        Args:
            transactions: List of Transaction objects.

        Returns:
            DataFrame with transaction data.
        """
        data = [
            {
                "Date": txn.date,
                "Description": txn.name,
                "Amount": txn.amount,
                "Category": txn.category,
                "Account": txn.spent_from,
                "Notes": txn.notes,
                "Transaction ID": txn.transaction_id,
            }
            for txn in transactions
        ]
        return pd.DataFrame(data)

    # TODO: Add Google Sheets export
    # @staticmethod
    # def to_google_sheets(
    #     transactions: List[Transaction],
    #     spreadsheet_id: str,
    #     sheet_name: str = "Transactions",
    # ) -> None:
    #     """
    #     Export transactions to Google Sheets.
    #
    #     Requires gspread and google-auth-oauthlib setup.
    #
    #     Args:
    #         transactions: List of Transaction objects.
    #         spreadsheet_id: Google Sheets spreadsheet ID.
    #         sheet_name: Name of sheet to write to.
    #     """
    #     # Implementation with gspread
    #     pass
