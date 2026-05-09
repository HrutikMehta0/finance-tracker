import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from finance_tracker.models import Transaction
from finance_tracker.exporter import Exporter
from finance_tracker.deduplicator import generate_transaction_hash


def create_workbook_with_txid(path: Path, headers, initial_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(headers)
    for row in initial_rows:
        ws.append(row)
    wb.save(path)


def test_append_only_new_transactions(tmp_path):
    # Setup initial workbook with one transaction ID
    wb_path = tmp_path / "finance.xlsx"
    headers = ["Date", "Transaction Name", "Category", "Spent From", "Amount", "Notes", "Transaction ID"]
    initial_rows = [["2026-01-01", "Coffee Shop", "Food", "Chequing", 3.5, "", "tx-abc123"]]
    create_workbook_with_txid(wb_path, headers, initial_rows)

    # Create transactions: one duplicate, one new
    t1 = Transaction(date="2026-01-01", name="Coffee Shop", category="Food", spent_from="Chequing", amount=3.5, notes="", transaction_id="tx-abc123")
    t2 = Transaction(date="2026-02-01", name="Bookstore", category="Books", spent_from="Chequing", amount=12.99, notes="", transaction_id="tx-new-1")

    # Append using Exporter
    Exporter.to_excel([t1, t2], str(wb_path))

    wb = load_workbook(wb_path)
    ws = wb["Transactions"]
    # Header + initial + new = 3 rows
    assert ws.max_row == 3

    # Check that the new transaction id exists in sheet
    ids = [cell[0].value for cell in ws.iter_rows(min_row=2, min_col=7, max_col=7)]
    assert "tx-new-1" in ids

    # Running again should not add duplicates
    Exporter.to_excel([t1, t2], str(wb_path))
    wb2 = load_workbook(wb_path)
    ws2 = wb2["Transactions"]
    assert ws2.max_row == 3


def test_filter_new_transactions():
    t1 = Transaction(date="2026-01-01", name="A", category="", spent_from="", amount=1.0, notes="", transaction_id="id1")
    t2 = Transaction(date="2026-01-02", name="B", category="", spent_from="", amount=2.0, notes="", transaction_id="id2")
    existing = {"id1"}
    new = Exporter.filter_new_transactions([t1, t2], existing)
    assert len(new) == 1
    assert new[0].transaction_id == "id2"


def test_append_with_missing_txid_column(tmp_path):
    # Create workbook without Transaction ID column but with an initial transaction
    wb_path = tmp_path / "finance2.xlsx"
    headers = ["Date", "Transaction Name", "Category", "Spent From", "Amount", "Notes"]
    initial_rows = [["2026-01-01", "Coffee Shop", "Food", "Chequing", 3.50, ""]]
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(headers)
    for row in initial_rows:
        ws.append(row)
    wb.save(wb_path)

    # Compute expected transaction id for existing row
    existing_hash = generate_transaction_hash("2026-01-01", "Coffee Shop", 3.5, "Chequing")

    # Create incoming transactions: duplicate and new
    t1 = Transaction(date="2026-01-01", name="Coffee Shop", category="Food", spent_from="Chequing", amount=3.5, notes="", transaction_id=existing_hash)
    t2 = Transaction(date="2026-02-02", name="New Shop", category="Other", spent_from="Chequing", amount=5.0, notes="", transaction_id="tx-new-2")

    Exporter.to_excel([t1, t2], str(wb_path))

    wb2 = load_workbook(wb_path)
    ws2 = wb2["Transactions"]
    # Header + initial + new = 3 rows
    assert ws2.max_row == 3
